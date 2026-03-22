"""Collect all Writing student data and generate dashboard JSON."""

import argparse
import json
import logging
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

# Add parent dir to path so we can import writing_automation
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from writing_automation.api_client import TimebackAPI
from writing_automation.config import (
    GRADE_SEQUENCES,
    MINUTES_GOAL_PER_DAY,
    PASS_THRESHOLD,
    RUSH_THRESHOLD,
    SESSIONS,
    XP_GOAL_PER_DAY,
)
from writing_automation.csv_loader import load_csv
from writing_automation.deep_dive import detect_deep_dives
from writing_automation.deep_dive_analysis import (
    analyze_with_claude,
    fetch_and_parse_tests,
    identify_deep_dive_tests,
    is_rushed,
    _get_anthropic_client,
)
from writing_automation.enrollment_fetcher import (
    fetch_student_profiles,
    fetch_writing_enrollments,
)
from writing_automation.hmg_calculator import compute_all_hmg
from writing_automation.student_progress import _get_level
from writing_automation.student_progress import _count_weekdays
from writing_automation.test_type_mapper import classify_test_types
# XP is now computed per-student from raw activity results (not the bulk fetcher)

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

import re as _re

GRADEBOOK_BASE = "/ims/oneroster/gradebook/v1p2"

_UUID_RE = _re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", _re.I)


def _is_uuid(s: str) -> bool:
    return bool(_UUID_RE.match(s))


def _is_alphawrite(ali_sid: str) -> bool:
    """Check if an assessmentLineItem sourcedId is an AlphaWrite activity.

    AlphaWrite activities use two prefix formats:
    - 'alphawrite-' (sentences, paragraphs, etc.)
    - 'alphawrite:' (compositions/essays)
    """
    return ali_sid.startswith("alphawrite-") or ali_sid.startswith("alphawrite:")


ACCURACY_THRESHOLD = 80
OUTPUT_PATH = Path(__file__).resolve().parent / "data.json"

# Default paths for EG and S1 snapshot data
DEFAULT_EG_CSV = Path(__file__).resolve().parent.parent / "Student_Progress_Tra_1773079782808.csv"
DEFAULT_S1_SNAPSHOT = Path(__file__).resolve().parent.parent / "SY25-26 Session 1 Snapshot (Academics).xlsx"

# Nickname -> full name mappings for S1 snapshot matching
_S1_NAME_OVERRIDES = {
    "abi constain": "abigail constain",
    "benny valles": "benjamin valles",
    "bobbi brown": "bobbi sue brown",
    "cami fernandez": "camila fernandez",
    "dario poyatos ramos": "dario ramos",
    "daveyp paul": "david paul",
    "dax hummel": "daxon hummel",
    "des pardi": "desmond pardi",
    "izzy vicente": "isabella vicente",
    "ju orloff": "juliana orloff",
    "nathan scharf": "nathaniel scharf",
    "penny marty": "penelope marty",
    "saeed tarawneh": "said tarawneh",
    "sebi cobas": "sebastian cobas",
    "bella barba": "isabella barba",
    "ben de amorim": "ben deamorim",
    "gus haig": "august haig",
    "grey walker": "greyson walker",
}


def load_s1_writing_names(snapshot_path: str) -> set[str]:
    """Load S1 Writing-enrolled student names from the S1 Snapshot spreadsheet.

    Uses the XPschoolday sheet — students with Writing XP > 0 in S1.
    Returns a set of lowercase names, with nickname overrides applied.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(snapshot_path), data_only=True)
    ws = wb["XPschoolday"]
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        fullname = row[2]
        subject = row[3]
        xp = row[4]
        if not fullname or not isinstance(fullname, str):
            continue
        if str(subject).strip() == "Writing" and isinstance(xp, (int, float)) and xp > 0:
            raw = fullname.strip().lower()
            names.add(_S1_NAME_OVERRIDES.get(raw, raw))
    wb.close()
    logger.info("Loaded %d S1 Writing students from snapshot", len(names))
    return names

# ---------------------------------------------------------------------------
# Session cookie helpers (interactive prompting on expiry)
# ---------------------------------------------------------------------------

def _load_session_cookie() -> str:
    """Load the Alpha session cookie from .env, prompting if missing."""
    from dotenv import load_dotenv
    from writing_automation.config import ALPHA_SESSION_COOKIE_ENV, ENV_FILE

    load_dotenv(ENV_FILE)
    cookie = os.getenv(ALPHA_SESSION_COOKIE_ENV, "")
    if not cookie:
        cookie = _prompt_for_cookie(ENV_FILE, ALPHA_SESSION_COOKIE_ENV)
    return cookie


def _prompt_for_cookie(env_file, env_key: str) -> str:
    """Prompt the user to paste a new session cookie and save it to .env."""
    print("\n" + "=" * 60)
    print("Session cookie required for test page fetching.")
    print("Open Alpha in your browser, copy the 'session' cookie value,")
    print("and paste it below.")
    print("=" * 60)
    cookie = input("Session cookie: ").strip()
    if not cookie:
        raise ValueError("No cookie provided. Cannot fetch test pages.")
    _save_cookie_to_env(env_file, env_key, cookie)
    os.environ[env_key] = cookie
    return cookie


def _save_cookie_to_env(env_file, env_key: str, cookie: str):
    """Update or add the cookie in the .env file."""
    env_path = Path(env_file)
    if env_path.exists():
        content = env_path.read_text(encoding="utf-8")
        # Replace existing line or append
        import re as _re2
        pattern = _re2.compile(rf'^{_re2.escape(env_key)}=.*$', _re2.MULTILINE)
        if pattern.search(content):
            content = pattern.sub(f'{env_key}={cookie}', content)
        else:
            content = content.rstrip() + f'\n{env_key}={cookie}\n'
        env_path.write_text(content, encoding="utf-8")
    else:
        env_path.write_text(f'{env_key}={cookie}\n', encoding="utf-8")
    logger.info("Session cookie saved to %s", env_path)


def _fetch_and_parse_with_retry(test_results, session_cookie: str) -> tuple[list[dict], str]:
    """Fetch and parse tests, prompting for a new cookie on auth failure.

    Returns (parsed_tests, final_cookie).
    """
    from writing_automation.config import ALPHA_SESSION_COOKIE_ENV, ENV_FILE

    try:
        parsed = fetch_and_parse_tests(test_results, session_cookie)
        return parsed, session_cookie
    except ValueError as e:
        if "cookie expired" in str(e).lower():
            logger.warning("Session cookie expired. Prompting for a new one...")
            new_cookie = _prompt_for_cookie(ENV_FILE, ALPHA_SESSION_COOKIE_ENV)
            parsed = fetch_and_parse_tests(test_results, new_cookie)
            return parsed, new_cookie
        raise


def run_deep_dive_analysis(
    deep_dive_tests: dict[tuple[str, int], list],
    email_to_name: dict[str, str],
) -> dict[tuple[str, int], dict]:
    """Run Claude analysis for all deep dive student/grade combos.

    Returns dict mapping (email, grade) -> analysis dict with keys:
    questions_missed, error_analysis, root_causes, recommended_actions
    """
    if not deep_dive_tests:
        return {}

    total_tests = sum(len(v) for v in deep_dive_tests.values())
    logger.info(
        "Deep Dive Analysis: %d student/grade combos, %d tests to analyze",
        len(deep_dive_tests), total_tests,
    )

    # Load cookie
    session_cookie = _load_session_cookie()

    # Init Claude client
    client = _get_anthropic_client()

    analyses = {}
    for i, ((email, grade), test_list) in enumerate(sorted(deep_dive_tests.items()), 1):
        name = email_to_name.get(email, email)
        logger.info(
            "  [%d/%d] Analyzing %s at G%d (%d tests)...",
            i, len(deep_dive_tests), name, grade, len(test_list),
        )

        # Build rushing info
        rushing_info = []
        for r in test_list:
            rushed = is_rushed(r.time_spent_seconds, grade)
            rushing_info.append({
                "test_name": r.test_name,
                "time_seconds": r.time_spent_seconds,
                "rushed": rushed,
                "score": r.score,
                "date": r.score_date,
            })

        # Fetch and parse test pages (with cookie retry)
        try:
            parsed_tests, session_cookie = _fetch_and_parse_with_retry(
                test_list, session_cookie
            )
        except ValueError as e:
            logger.error("  Cannot fetch tests for %s G%d: %s", name, grade, e)
            analyses[(email, grade)] = {
                "questions_missed": "",
                "error_analysis": f"Could not fetch test pages: {e}",
                "root_causes": "",
                "recommended_actions": "",
            }
            continue

        if not parsed_tests:
            logger.warning("  No tests could be parsed for %s G%d", name, grade)
            analyses[(email, grade)] = {
                "questions_missed": "",
                "error_analysis": "Could not fetch/parse test pages",
                "root_causes": "",
                "recommended_actions": "",
            }
            continue

        # Analyze with Claude
        try:
            analysis = analyze_with_claude(client, name, grade, parsed_tests, rushing_info)
            analyses[(email, grade)] = analysis
            logger.info("  Analysis complete for %s G%d", name, grade)
        except Exception as e:
            logger.error("  Claude analysis failed for %s G%d: %s", name, grade, e)
            analyses[(email, grade)] = {
                "questions_missed": "",
                "error_analysis": f"Analysis failed: {e}",
                "root_causes": "",
                "recommended_actions": "",
            }

    return analyses


def load_effective_grades(csv_path: str) -> tuple[dict[str, int], dict[str, int]]:
    """Load effective grades from the Student Progress Tracker CSV.

    The CSV may contain rows for multiple subjects. We separate Writing
    and Language EGs.

    Returns (writing_eg_by_name, language_eg_by_name) both mapping
    student name (lowercase) -> effective grade (int).
    """
    import csv as _csv

    writing_eg: dict[str, int] = {}
    language_eg: dict[str, int] = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in _csv.DictReader(f):
            name = row.get("Student", "").strip()
            eg = row.get("Effective Grade", "").strip()
            subject = row.get("Subject", "Writing").strip()
            if not name or not eg:
                continue
            try:
                eg_val = int(eg)
            except ValueError:
                continue
            key = name.lower()
            if subject == "Language":
                language_eg[key] = eg_val
            else:
                writing_eg[key] = eg_val

    logger.info("Loaded effective grades: %d Writing, %d Language from %s",
                len(writing_eg), len(language_eg), csv_path)
    return writing_eg, language_eg


def _school_days_to_date(session_name: str) -> int:
    """Count school days from session school_start to yesterday.

    Because the dashboard is always updated the following day (due to timezone
    differences), we use ``today - 1 day`` as the cutoff so we don't
    under-track students.
    """
    session = SESSIONS[session_name]
    start = datetime.strptime(session.get("school_start", session["start"]), "%Y-%m-%d")
    end = datetime.strptime(session["end"], "%Y-%m-%d")
    yesterday = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
    cutoff = min(yesterday, end)
    if cutoff < start:
        return 0
    return _count_weekdays(start, cutoff)

# ---------------------------------------------------------------------------
# A&D Master Roster — used as whitelist and source of truth for campus
# ---------------------------------------------------------------------------

ROSTER_PATH = Path(__file__).resolve().parent.parent / "A&D Master Roster 25-26 - Master.csv"

# Student Group values that should be excluded
_EXCLUDED_GROUPS = {"mock", "mock student", "shadow", "test", "guide"}

# Individual student emails to exclude
_EXCLUDED_EMAILS = {
    "lincoln.thomas@alpha.school",
    "luka.scaletta@alpha.school",
    "elle.liemandt@alpha.school",
}

# Students whose roster status should be treated as "Enrolled" (e.g. transfers)
_STATUS_OVERRIDES = {
    "quinn.oneal@2hourlearning.com",
    "robin.oneal@2hourlearning.com",
    "atlas.kloiber@alpha.school",
    "lincoln.kloiber@alpha.school",
    "eva.quintero@2hourlearning.com",
    "scarlett.oneal@2hourlearning.com",
}

# Legacy Dash campuses (case-insensitive matching via _normalise)
_LEGACY_CAMPUSES = {
    "alpha anywhere (homeschool)",
    "novatio",
    "unbound academy",
    "kairos learning solutions",
    "lipscomb academy accelerate",
}

# Campuses excluded from the Timeback page (but not Legacy Dash)
_TIMEBACK_EXCLUDED_CAMPUSES = {
    "2 hour learning",
    "2 hour single user",
    "alpha k-8",
    "aie elite prep",
    "alpha austin 25' ai summer camp",
    "alpha international test school",
    "alphalearn",
    "beyond ai",
    "guide school",
    "high school sat prep",
    "mock school org",
    "speedrun",
    "school in the hills",
    "trilogy central support",
    "centner academy",
    "colearn academy",
    "the st. james performance academy",
}


def _load_roster() -> dict[str, dict]:
    """Load A&D Master Roster. Returns {email_lower: {campus, level, grade, group, name}}."""
    import csv
    roster: dict[str, dict] = {}
    if not ROSTER_PATH.exists():
        logger.warning("A&D Master Roster not found at %s", ROSTER_PATH)
        return roster
    with open(ROSTER_PATH, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            email = row.get("Student Alpha Email", "").strip().lower()
            status = row.get("Admission Status", "").strip()
            group = row.get("Student Group", "").strip().lower()
            campus = row.get("Campus", "").strip()
            if not email or (status != "Enrolled" and email not in _STATUS_OVERRIDES):
                continue
            if group in _EXCLUDED_GROUPS:
                continue
            roster[email] = {
                "campus": campus,
                "level": row.get("Current Level", "").strip(),
                "grade": row.get("Current Grade Level", "").strip(),
                "name": row.get("Full Name", "").strip(),
                "group": group,
            }
    logger.info("Loaded %d enrolled students from A&D Master Roster", len(roster))
    return roster


def _classify_dashboard(campus: str) -> str:
    """Return 'legacy' or 'timeback' for a campus, or '' if excluded."""
    c = campus.lower()
    if c in _LEGACY_CAMPUSES:
        return "legacy"
    if c in _TIMEBACK_EXCLUDED_CAMPUSES:
        return ""
    return "timeback"


# ---------------------------------------------------------------------------
# AlphaWrite skill plan name mapping
# ---------------------------------------------------------------------------

def _load_skill_plan() -> dict[str, tuple[str, str]]:
    """Load AlphaWrite skill plan from xlsx and return {skill_id: (name, course)} mapping."""
    skill_plan_path = Path(__file__).resolve().parent.parent / "AlphaWrite Skill Plan 2025_2026 (1).xlsx"
    if not skill_plan_path.exists():
        logger.warning("AlphaWrite Skill Plan not found at %s", skill_plan_path)
        return {}

    import openpyxl
    wb = openpyxl.load_workbook(str(skill_plan_path), read_only=True)
    ws = wb["new-aw-skill-plan 2025"]
    mapping: dict[str, tuple[str, str]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        skill_id, name, qs_id = row[0], row[1], row[2]
        course = row[6] if len(row) > 6 else ""
        if skill_id and name:
            mapping[str(skill_id).lower()] = (str(name), str(course or ""))
        if qs_id and name:
            # Also map the QS_ID suffix (e.g. FRAGMENT_OR_SENTENCE from 3.2.FRAGMENT_OR_SENTENCE)
            qs_str = str(qs_id)
            mapping[qs_str.lower()] = (str(name), str(course or ""))
            if "." in qs_str:
                suffix = qs_str.rsplit(".", 1)[-1].lower()
                # Don't overwrite if already exists with a different name
                if suffix not in mapping:
                    mapping[suffix] = (str(name), str(course or ""))
    wb.close()
    logger.info("Loaded %d skill plan mappings", len(mapping))
    return mapping


_SKILL_PLAN: dict[str, tuple[str, str]] | None = None


def _get_skill_plan() -> dict[str, tuple[str, str]]:
    global _SKILL_PLAN
    if _SKILL_PLAN is None:
        _SKILL_PLAN = _load_skill_plan()
    return _SKILL_PLAN


def _resolve_compositions_name(ali_sid: str) -> tuple[str, str]:
    """Resolve an alphawrite:compositions: activity to (readable_name, course).

    Example ali_sid: alphawrite:compositions:essays-g6:essay-3-tfxy:stage-revise:
    Returns e.g. ("Revise (Essay 3)", "Essays G6")
    """
    parts = ali_sid.split(":")
    # parts: ['alphawrite', 'compositions', 'essays-g6', 'essay-3-tfxy', 'stage-revise', '']
    course = ""
    stage = ""
    essay_label = ""
    if len(parts) >= 3:
        # e.g. 'essays-g6' -> 'Essays G6'
        course = parts[2].replace("-", " ").title()
    if len(parts) >= 4:
        # e.g. 'essay-3-tfxy' -> 'Essay 3'
        essay_part = parts[3]
        m = _re.match(r"essay-(\d+)", essay_part)
        if m:
            essay_label = f"Essay {m.group(1)}"
    if len(parts) >= 5:
        # e.g. 'stage-revise' or 'stage-draft:attempt-1'
        stage_part = parts[4]
        stage = stage_part.replace("stage-", "").replace("-", " ").title()
        if stage.startswith("Identif"):
            stage = "Setup"  # stage-identify maps to "Setup" in the admin

    name = stage if stage else "Compositions Activity"
    if essay_label:
        name = f"{name} ({essay_label})"

    return (name, course)


def _resolve_activity_name(ali_sid: str, meta: dict) -> tuple[str, str] | None:
    """Resolve an AlphaWrite activity to (readable_name, course).
    Returns None if the activity is not an AlphaWrite activity."""
    # Only process AlphaWrite activities
    if not _is_alphawrite(ali_sid):
        return None

    # Handle alphawrite:compositions: prefix (Essays)
    if ali_sid.startswith("alphawrite:compositions:"):
        return _resolve_compositions_name(ali_sid)

    skill_plan = _get_skill_plan()

    # Try direct lookup by full ali_sid (minus -assessment-line-item suffix)
    clean_sid = ali_sid.replace("-assessment-line-item", "").lower()
    if clean_sid in skill_plan:
        return skill_plan[clean_sid]

    # Try extracting the activity slug from the ali_sid
    # e.g. alphawrite-sentences-iii-identify-sentence-type -> identify-sentence-type
    parts = clean_sid.split("-", 3)  # ['alphawrite', 'sentences', 'iii', 'identify-sentence-type']
    if len(parts) >= 4:
        slug = parts[3].replace("-", "_").upper()
        if slug.lower() in skill_plan:
            return skill_plan[slug.lower()]

    # Try metadata activity field
    activity_name = meta.get("activity", "")
    if activity_name:
        if activity_name.lower() in skill_plan:
            return skill_plan[activity_name.lower()]
        # Try converting underscores
        slug = activity_name.lower().replace(" ", "_")
        if slug in skill_plan:
            return skill_plan[slug]

    # Fallback: derive readable name from ali_sid
    readable = clean_sid.replace("alphawrite-", "").replace("-", " ").title()
    # Derive course from the structure
    course = ""
    if "sentences-i-" in ali_sid and "sentences-ii" not in ali_sid and "sentences-iii" not in ali_sid:
        course = "Sentences G3"
    elif "sentences-ii-" in ali_sid and "sentences-iii" not in ali_sid:
        course = "Sentences G4"
    elif "sentences-iii" in ali_sid:
        course = "Sentences G5"
    elif "sentences-iv" in ali_sid:
        course = "Sentences G6"
    elif "sentences-v-" in ali_sid:
        course = "Sentences G7"
    elif "sentences-vi" in ali_sid:
        course = "Sentences G8"
    elif "paragraphs-g3" in ali_sid:
        course = "Paragraphs G3"
    elif "paragraphs-g4" in ali_sid:
        course = "Paragraphs G4"
    elif "paragraphs-g5" in ali_sid:
        course = "Paragraphs G5"
    elif "paragraphs-g6" in ali_sid:
        course = "Paragraphs G6"
    elif "paragraphs-g7" in ali_sid:
        course = "Paragraphs G7"
    elif "paragraphs-g8" in ali_sid:
        course = "Paragraphs G8"
    elif "compositions" in ali_sid or "essays" in ali_sid:
        if "g6" in ali_sid:
            course = "Essays G6"
        elif "g7" in ali_sid:
            course = "Essays G7"
        elif "g8" in ali_sid:
            course = "Essays G8"

    return (readable, course)


# ---------------------------------------------------------------------------
# New data fetching functions
# ---------------------------------------------------------------------------

_WRITING_TEST_RE = _re.compile(r"Writing\s+G\d|Alpha\s+Standardized\s+Writing", _re.I)
_GRADE_SUB_RE = _re.compile(r"G(\d+)\.(\d+)")


def _classify_unknown_test_types(tests: list[dict]) -> list[dict]:
    """Retroactively classify tests with empty test_type.

    Logic:
    - First ever test at G3.1 → 'placement'
    - .1 at a new grade after passing the previous grade → 'test out'
    - All other tests within a grade → 'end of course'
    """
    if not tests:
        return tests

    passed_grades: set[int] = set()
    first_test_seen = False

    for t in tests:  # already sorted by date
        if t.get("test_type"):
            # Track passed grades from known-type tests too
            m = _GRADE_SUB_RE.search(t.get("name", ""))
            if m and t.get("passed"):
                passed_grades.add(int(m.group(1)))
            first_test_seen = True
            continue

        m = _GRADE_SUB_RE.search(t.get("name", ""))
        if not m:
            t["test_type"] = "end of course"
            first_test_seen = True
            continue

        grade = int(m.group(1))
        sub = int(m.group(2))

        if not first_test_seen and sub == 1 and grade == 3:
            t["test_type"] = "placement"
        elif sub == 1 and (grade - 1) in passed_grades:
            t["test_type"] = "test out"
        else:
            t["test_type"] = "end of course"

        if t.get("passed"):
            passed_grades.add(grade)
        first_test_seen = True

    return tests


def fetch_writing_test_results(
    api: TimebackAPI, student_id: str
) -> list[dict]:
    """Fetch standardized writing test results from the API for a student.

    Uses a wider search (no subject filter) and filters client-side,
    so tests with missing metadata.subject are still captured.
    """
    try:
        data = api.get(
            f"{GRADEBOOK_BASE}/assessmentResults/",
            {
                "limit": 3000,
                "filter": (
                    f"student.sourcedId='{student_id}'"
                    " AND metadata.resultType='assessment'"
                ),
            },
        )
        results = []
        for r in data.get("assessmentResults", []):
            meta = r.get("metadata", {})
            subject = meta.get("subject", "")
            test_name = meta.get("testName", "")
            # Include if subject is Writing OR testName matches writing test pattern
            if subject != "Writing" and not _WRITING_TEST_RE.search(test_name):
                continue
            results.append({
                "name": test_name,
                "test_type": meta.get("testType", ""),
                "score": r.get("score"),
                "date": (r.get("scoreDate") or "")[:10],
                "assigned_at": (meta.get("assignedAt") or "")[:10],
                "assignment_id": meta.get("assignmentId"),
                "test_link": meta.get("testLink", ""),
                "total_questions": meta.get("totalQuestions"),
                "correct_questions": meta.get("correctQuestions"),
                "passed": (r.get("score") or 0) >= PASS_THRESHOLD,
            })
        results.sort(key=lambda x: x["date"])
        return _classify_unknown_test_types(results)
    except Exception as e:
        logger.warning("Failed to fetch test results for %s: %s", student_id, e)
        return []


def fetch_activity_results(
    api: TimebackAPI, student_id: str, session_start: str, session_end: str
) -> list[dict]:
    """Fetch per-activity assessment results for a student within a session."""
    try:
        data = api.get(
            f"{GRADEBOOK_BASE}/assessmentResults/",
            {
                "limit": 3000,
                "filter": (
                    f"student.sourcedId='{student_id}'"
                    f" AND scoreDate>='{session_start}'"
                    f" AND scoreDate<='{session_end}'"
                ),
            },
        )
        return data.get("assessmentResults", [])
    except Exception as e:
        logger.warning("Failed to fetch activities for %s: %s", student_id, e)
        return []


def extract_low_accuracy_activities(raw_results: list[dict]) -> list[dict]:
    """Extract AlphaWrite activities below the accuracy threshold."""
    low = []
    for r in raw_results:
        ali_sid = r.get("assessmentLineItem", {}).get("sourcedId", "")
        meta = r.get("metadata", {})

        # Only include AlphaWrite activities
        resolved = _resolve_activity_name(ali_sid, meta)
        if resolved is None:
            continue

        name, course = resolved

        accuracy = meta.get("accuracy")
        if accuracy is None:
            total = meta.get("totalQuestions")
            correct = meta.get("correctQuestions")
            if total and correct is not None and total > 0:
                accuracy = round(100 * correct / total)
            else:
                continue

        if accuracy >= ACCURACY_THRESHOLD:
            continue

        total_q = meta.get("totalQuestions", 0)
        correct_q = meta.get("correctQuestions", 0)

        low.append({
            "name": name,
            "course": course,
            "accuracy": accuracy,
            "questions": f"{correct_q}/{total_q}" if total_q else "?",
            "xp": meta.get("xp", 0),
            "attempt": meta.get("attemptNumber", meta.get("attempt", 1)),
            "date": (r.get("scoreDate") or "")[:10],
        })

    return low


def extract_repeated_activities(raw_results: list[dict]) -> list[dict]:
    """Find AlphaWrite activities where a student has multiple attempts."""
    by_activity: dict[str, list[dict]] = defaultdict(list)
    activity_names: dict[str, tuple[str, str]] = {}

    for r in raw_results:
        ali_sid = r.get("assessmentLineItem", {}).get("sourcedId", "")
        if not ali_sid:
            continue
        meta = r.get("metadata", {})

        # Only include AlphaWrite activities
        resolved = _resolve_activity_name(ali_sid, meta)
        if resolved is None:
            continue

        if ali_sid not in activity_names:
            activity_names[ali_sid] = resolved

        accuracy = meta.get("accuracy")
        if accuracy is None:
            total = meta.get("totalQuestions")
            correct = meta.get("correctQuestions")
            if total and correct is not None and total > 0:
                accuracy = round(100 * correct / total)
        attempt = meta.get("attemptNumber", meta.get("attempt", 1))
        by_activity[ali_sid].append({
            "accuracy": accuracy,
            "attempt": attempt,
            "date": (r.get("scoreDate") or "")[:10],
        })

    repeated = []
    for ali_sid, attempts in by_activity.items():
        if len(attempts) <= 1:
            continue
        max_attempt = max(a.get("attempt", 1) or 1 for a in attempts)
        if max_attempt <= 1:
            continue

        accuracies = [a["accuracy"] for a in attempts if a["accuracy"] is not None]
        name, course = activity_names.get(ali_sid, (ali_sid, ""))

        repeated.append({
            "name": name,
            "course": course,
            "attempts": max_attempt,
            "best_accuracy": max(accuracies) if accuracies else None,
            "latest_accuracy": accuracies[-1] if accuracies else None,
        })

    return repeated


def compute_hmg_from_api_tests(api_tests: list[dict]) -> int:
    """Compute HMG from API test results.

    HMG = the grade of the highest Writing test the student has passed.
    Passing any test at a grade level means that grade is mastered.
    """
    hmg = 2  # Pre-G3 baseline
    for t in api_tests:
        if not t.get("passed"):
            continue
        m = _re.search(r"G(\d+)", t.get("name", ""))
        if m:
            grade = int(m.group(1))
            if grade > hmg:
                hmg = grade
    return hmg


def infer_next_test(hmg: int, test_history: list[dict]) -> dict | None:
    """Infer the next expected test based on HMG.

    When a student passes a test, they advance to the next grade level.
    So the next test is always G{HMG+1}.1.
    """
    if hmg >= 8:
        return None  # Completed all grades

    next_grade = hmg + 1
    if next_grade not in GRADE_SEQUENCES:
        return None

    test_name = f"G{next_grade}.1"

    # Check if student has already attempted a test at this grade level
    taken = any(
        f"G{next_grade}." in t.get("name", "")
        for t in test_history
        if not t.get("passed")
    )

    return {
        "name": test_name,
        "reason": f"HMG is G{hmg}, next grade level is G{next_grade}",
        "status": "retaking" if taken else "pending",
    }


def extract_xp_and_details(raw_results: list[dict]) -> dict:
    """Extract per-activity and per-test XP breakdowns, plus compute XP totals.

    A result is counted as Writing XP if:
    - metadata.subject == 'Writing', OR
    - assessmentLineItem.sourcedId is an AlphaWrite activity (alphawrite- or alphawrite:)

    Returns dict with 'activity_xp', 'test_xp' lists, and XP totals.
    """
    activity_xp_items = []
    test_xp_items = []
    alphawrite_xp_total = 0.0
    mastery_track_xp_total = 0.0
    test_xp_total = 0.0

    for r in raw_results:
        meta = r.get("metadata", {})
        xp = meta.get("xp", 0) or 0
        if xp <= 0:
            continue

        ali_sid = r.get("assessmentLineItem", {}).get("sourcedId", "")
        subject = meta.get("subject", "")
        is_alphawrite = _is_alphawrite(ali_sid)

        # Only include Writing-subject OR AlphaWrite activities
        if subject != "Writing" and not is_alphawrite:
            continue

        result_type = meta.get("resultType", "")
        lesson_type = meta.get("lessonType", "")
        date = (r.get("scoreDate") or "")[:10]

        if result_type == "assessment":
            # This is a writing test
            test_xp_total += xp
            test_xp_items.append({
                "name": meta.get("testName", "Unknown Test"),
                "xp": xp,
                "score": r.get("score"),
                "date": date,
            })
        else:
            # Activity XP
            resolved = _resolve_activity_name(ali_sid, meta)
            if resolved:
                name, course = resolved
                alphawrite_xp_total += xp
            elif is_alphawrite:
                # Fallback: derive readable name from ali_sid
                clean = ali_sid.replace("alphawrite-", "").replace("alphawrite:", "")
                name = clean.replace("-assessment-line-item", "").replace("-", " ").replace(":", " ").title()
                course = ""
                alphawrite_xp_total += xp
            else:
                # Non-AlphaWrite Writing activity (mastery track / external lesson)
                app_name = meta.get("appName", "")
                test_name = meta.get("testName", "")
                name = test_name or app_name or "Writing Activity"
                course = ""
                if name.startswith("caliper_") or name.startswith("Caliper_"):
                    name = app_name or "Mastery Track Activity"
                elif name.startswith("Nice_"):
                    name = name.replace("Nice_", "").replace("_", " ").title()
                elif _is_uuid(name):
                    name = app_name or "Writing Activity"
                mastery_track_xp_total += xp

            activity_xp_items.append({
                "name": name,
                "course": course,
                "xp": xp,
                "date": date,
                "type": "alphawrite" if (is_alphawrite or lesson_type == "powerpath-100") else (
                    "external" if lesson_type == "external-lesson" else "mastery_track"
                ),
            })

    # Sort by date
    activity_xp_items.sort(key=lambda x: x["date"])
    test_xp_items.sort(key=lambda x: x["date"])

    # Determine most recent XP date
    all_dates = [a["date"] for a in activity_xp_items if a["date"]] + \
                [t["date"] for t in test_xp_items if t["date"]]
    last_xp_date = max(all_dates) if all_dates else None

    return {
        "activity_xp": activity_xp_items,
        "test_xp": test_xp_items,
        "alphawrite_xp": alphawrite_xp_total,
        "mastery_track_xp": mastery_track_xp_total,
        "test_xp_total": test_xp_total,
        "last_xp_date": last_xp_date,
    }


# ---------------------------------------------------------------------------
# Main collector
# ---------------------------------------------------------------------------

def collect(csv_path: str, session_name: str, *, skip_analysis: bool = False, effective_grades_csv: str | None = None, s1_snapshot_path: str | None = None) -> dict:
    """Collect all data and return the dashboard JSON structure."""
    session = SESSIONS[session_name]
    session_start = session["start"]
    session_end = session["end"]
    school_days = _school_days_to_date(session_name)

    # 1. Load A&D Master Roster (whitelist + campus source of truth)
    roster = _load_roster()

    # 2. Load CSV
    logger.info("Loading CSV: %s", csv_path)
    csv_results = load_csv(csv_path)
    logger.info("Loaded %d CSV results", len(csv_results))

    # 3. Init API
    logger.info("Initializing Timeback API...")
    api = TimebackAPI()

    # 4. Fetch enrollments + profiles
    logger.info("Fetching Writing enrollments...")
    enrollments = fetch_writing_enrollments(api)
    student_ids = set(enrollments.keys())
    logger.info("Found %d enrolled students", len(student_ids))

    logger.info("Fetching student profiles...")
    profiles = fetch_student_profiles(api, student_ids)

    # 4. Compute HMG + classify
    hmg_map = compute_all_hmg(csv_results)
    classifications = classify_test_types(csv_results)

    # 6. Detect deep dives
    deep_dives = detect_deep_dives(csv_results)
    deep_dive_tests = identify_deep_dive_tests(csv_results, deep_dives, session_name)

    # 6b. Run Claude deep dive analysis for students in testing loops
    email_to_name = {}
    for r in csv_results:
        email_to_name[r.student_email] = r.student_name
    if skip_analysis:
        logger.info("Skipping Claude deep dive analysis (--skip-analysis)")
        dd_analyses = {}
    else:
        dd_analyses = run_deep_dive_analysis(deep_dive_tests, email_to_name)

    # 6c. Load effective grades from CSV
    eg_by_name: dict[str, int] = {}
    lang_eg_by_name: dict[str, int] = {}
    if effective_grades_csv:
        eg_by_name, lang_eg_by_name = load_effective_grades(effective_grades_csv)

    # 6d. Load S1 cohort names
    s1_names: set[str] = set()
    if s1_snapshot_path:
        s1_names = load_s1_writing_names(s1_snapshot_path)

    # 7. Build email -> csv results map
    csv_by_email: dict[str, list] = defaultdict(list)
    for r in csv_results:
        csv_by_email[r.student_email].append(r)

    # 8. Goals
    xp_goal = XP_GOAL_PER_DAY * school_days
    minutes_goal = MINUTES_GOAL_PER_DAY * school_days

    # 9. Assemble per-student data
    students = []
    total = len(student_ids)
    for idx, sid in enumerate(sorted(student_ids), 1):
        profile = profiles.get(sid)
        if not profile:
            continue

        email = profile.email

        # Only include students in the A&D Master Roster
        roster_entry = roster.get(email.lower())
        if not roster_entry:
            continue

        # Skip individually excluded students
        if email.lower() in _EXCLUDED_EMAILS:
            continue

        # Use roster campus as source of truth
        roster_campus = roster_entry["campus"]

        # Classify into dashboard group — only include Timeback students
        dash_group = _classify_dashboard(roster_campus)
        if dash_group != "timeback":
            continue

        logger.info("Processing student %d/%d: %s", idx, total, email)

        # Enrollments
        student_enrollments = enrollments.get(sid, [])

        # Level
        level = _get_level(profile.age_grade)

        # Fetch test history from API
        api_tests = fetch_writing_test_results(api, sid)

        # HMG — computed from API test results (highest grade with a passed test)
        hmg = compute_hmg_from_api_tests(api_tests)
        # Starting HMG from placement tests in CSV
        starting_hmg = 2
        student_csv = csv_by_email.get(email, [])
        if student_csv:
            placement = [r for r in student_csv if r.csv_test_type == "Placement"]
            if placement:
                passed = {r.test_grade for r in placement if r.score >= PASS_THRESHOLD}
                for g in range(3, 9):
                    if g in passed:
                        starting_hmg = g
                    else:
                        break

        # G8 completion check
        completed_g8 = hmg >= 8

        # Last test
        last_test = None
        if api_tests:
            t = api_tests[-1]
            last_test = {
                "name": t["name"],
                "type": t["test_type"],
                "score": t["score"],
                "date": t["date"],
                "passed": t["passed"],
            }

        # Session tests (from CSV for detailed info including time)
        session_start_dt = datetime.strptime(session_start, "%Y-%m-%d")
        session_end_dt = datetime.strptime(session_end, "%Y-%m-%d")
        session_tests = []
        for r in sorted(student_csv, key=lambda x: x.score_date):
            if session_start_dt <= r.score_date <= session_end_dt:
                key = (r.student_email, r.test_name, r.score_date)
                test_type = classifications.get(key, r.csv_test_type)
                session_tests.append({
                    "name": r.test_name,
                    "type": test_type,
                    "score": r.score,
                    "date": r.score_date.strftime("%Y-%m-%d"),
                    "passed": r.score >= PASS_THRESHOLD,
                    "time_seconds": r.time_spent_seconds,
                    "rushed": is_rushed(r.time_spent_seconds, r.test_grade),
                })

        # Next expected test (passing a test advances to next grade level)
        next_test = infer_next_test(hmg, api_tests)

        # Deep dive — only for grades NOT yet mastered (grade > hmg)
        dd_needed = any((email, g) in deep_dives for g in range(3, 9) if g > hmg)
        dd_details = []
        for (dd_email, dd_grade), dd_tests in deep_dive_tests.items():
            if dd_email != email:
                continue
            # Skip deep dives for grades the student has already mastered
            if dd_grade <= hmg:
                continue
            failed = [t for t in dd_tests if t.score < PASS_THRESHOLD]
            rushed_count = sum(1 for t in dd_tests if is_rushed(t.time_spent_seconds, dd_grade))
            times = [t.time_spent_seconds for t in dd_tests if t.time_spent_seconds]
            avg_time = round(sum(times) / len(times) / 60, 1) if times else 0

            dd_details.append({
                "grade": dd_grade,
                "total_tests": len(dd_tests),
                "failed_count": len(failed),
                "rushed_count": rushed_count,
                "avg_time_minutes": avg_time,
                "tests": [
                    {
                        "name": t.test_name,
                        "score": t.score,
                        "date": t.score_date.strftime("%Y-%m-%d"),
                        "rushed": is_rushed(t.time_spent_seconds, dd_grade),
                    }
                    for t in dd_tests
                ],
                "analysis": dd_analyses.get((dd_email, dd_grade)),
            })

        # Fetch activity results for accuracy analysis, XP details, and XP totals
        raw_activities = fetch_activity_results(api, sid, session_start, session_end)
        xp_result = extract_xp_and_details(raw_activities)
        xp_details = xp_result
        alphawrite_xp = xp_result["alphawrite_xp"]
        mastery_track_xp = xp_result["mastery_track_xp"]
        test_xp_val = xp_result["test_xp_total"]
        total_xp = alphawrite_xp + mastery_track_xp + test_xp_val

        # Split XP into school vs break periods for accurate goal tracking
        school_start_date = session.get("school_start", session_start)
        all_xp_items = xp_result.get("activity_xp", []) + xp_result.get("test_xp", [])
        school_xp = sum(a["xp"] for a in all_xp_items if a.get("date", "") >= school_start_date)
        break_xp = total_xp - school_xp
        avg_xp = round(school_xp / school_days, 1) if school_days else 0

        # For G8 completers, skip accuracy/deep dive/enrollment analysis
        if completed_g8:
            low_accuracy = []
            repeated = []
            enrollment_mismatch = None
            insights = []
        else:
            low_accuracy = extract_low_accuracy_activities(raw_activities)
            repeated = extract_repeated_activities(raw_activities)

            # Enrollment mismatch
            enrollment_mismatch = None
            if student_enrollments:
                expected = hmg + 1
                enrolled_grades = []
                for e in student_enrollments:
                    m = _re.search(r"G(\d+)", e)
                    if m:
                        enrolled_grades.append(int(m.group(1)))
                if enrolled_grades and expected not in enrolled_grades:
                    actual = ", ".join(f"G{g}" for g in enrolled_grades)
                    enrollment_mismatch = f"Expected G{expected}, enrolled in {actual}"

            # Build insights
            insights = []
            if dd_needed:
                for d in dd_details:
                    rushed_txt = f" ({d['rushed_count']} rushed)" if d["rushed_count"] else ""
                    insights.append({
                        "type": "deep_dive",
                        "severity": "high",
                        "text": f"Deep Dive at G{d['grade']}: {d['failed_count']} failed tests{rushed_txt}",
                    })
            if low_accuracy:
                insights.append({
                    "type": "low_accuracy",
                    "severity": "medium",
                    "text": f"{len(low_accuracy)} activit{'y' if len(low_accuracy) == 1 else 'ies'} below {ACCURACY_THRESHOLD}% accuracy",
                })
            if repeated:
                for rep in repeated[:3]:
                    insights.append({
                        "type": "repeated",
                        "severity": "low",
                        "text": f"Repeating '{rep['name']}' in {rep['course']} ({rep['attempts']} attempts)",
                    })
            if school_xp < xp_goal and school_days > 0:
                pct = round(100 * school_xp / xp_goal) if xp_goal > 0 else 0
                last_xp = xp_details.get("last_xp_date")
                xp_text = f"XP behind target: {round(school_xp)}/{round(xp_goal)} ({pct}%)"
                if last_xp:
                    xp_text += f" — last XP earned {last_xp}"
                insights.append({
                    "type": "goal_xp",
                    "severity": "medium",
                    "text": xp_text,
                })
            if enrollment_mismatch:
                insights.append({
                    "type": "enrollment_mismatch",
                    "severity": "medium",
                    "text": enrollment_mismatch,
                })

        # Test summary stats (from all-time api_tests)
        passed_tests = [t for t in api_tests if t["passed"]]
        test_summary = {
            "total_taken": len(api_tests),
            "total_passed": len(passed_tests),
            "end_of_course_passed": sum(1 for t in passed_tests if t["test_type"] == "end of course"),
            "test_outs_passed": sum(1 for t in passed_tests if t["test_type"] == "test out"),
            "placement_passed": sum(1 for t in passed_tests if t["test_type"] == "placement"),
        }

        students.append({
            "id": sid,
            "name": profile.full_name,
            "email": email,
            "campus": roster_campus,
            "dashboard": dash_group,
            "level": level,
            "age_grade": profile.age_grade,
            "hmg": hmg,
            "starting_hmg": starting_hmg,
            "grades_advanced": hmg - starting_hmg,
            "effective_grade": eg_by_name.get(profile.full_name.lower()),
            "effective_grades_mastered": max(0, hmg - (eg_by_name.get(profile.full_name.lower(), hmg + 1) - 1)) if eg_by_name.get(profile.full_name.lower()) else None,
            "language_eg": lang_eg_by_name.get(profile.full_name.lower()),
            "s1_cohort": profile.full_name.lower() in s1_names if s1_names else None,
            "completed_g8": completed_g8,
            "enrollments": student_enrollments,
            "still_enrolled": bool(student_enrollments),
            "last_test": last_test,
            "next_expected_test": next_test,
            "all_tests": api_tests,
            "test_summary": test_summary,
            "xp": {
                "alphawrite": round(alphawrite_xp, 1),
                "mastery_track": round(mastery_track_xp, 1),
                "test": round(test_xp_val, 1),
                "total": round(total_xp, 1),
                "school": round(school_xp, 1),
                "break": round(break_xp, 1),
                "goal_to_date": round(xp_goal, 1),
                "avg_per_day": avg_xp,
                "meets_goal": school_xp >= xp_goal,
                "last_xp_date": xp_details.get("last_xp_date"),
            },
            "xp_details": xp_details,
            "session_tests": session_tests,
            "accuracy": {
                "activities_below_threshold": low_accuracy,
                "repeated_activities": repeated,
            },
            "deep_dive": {
                "needed": dd_needed if not completed_g8 else False,
                "details": dd_details if not completed_g8 else [],
            },
            "insights": insights,
            "enrollment_mismatch": enrollment_mismatch,
        })

    # All session definitions for per-session metrics
    all_sessions = {
        "S1": {"start": "2025-08-11", "end": "2025-10-17", "label": "Session 1"},
        "S2": {"start": "2025-10-20", "end": "2026-01-02", "label": "Session 2"},
        "S3": {"start": "2026-01-05", "end": "2026-02-20", "label": "Session 3"},
        "S4": {"start": "2026-02-21", "end": "2026-04-17", "label": "Session 4"},
    }

    # Top-level structure
    dashboard = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "session": {
            "name": session_name,
            "start": session_start,
            "end": session_end,
            "school_start": session.get("school_start", session_start),
            "school_days_elapsed": school_days,
        },
        "all_sessions": all_sessions,
        "thresholds": {
            "xp_per_day": XP_GOAL_PER_DAY,
            "minutes_per_day": MINUTES_GOAL_PER_DAY,
            "accuracy_pct": ACCURACY_THRESHOLD,
            "pass_score": PASS_THRESHOLD,
        },
        "students": students,
    }

    return dashboard


def main():
    parser = argparse.ArgumentParser(description="Collect Writing dashboard data")
    parser.add_argument("csv", help="Path to writing-results CSV")
    parser.add_argument("--session", default="S4", choices=list(SESSIONS.keys()))
    parser.add_argument("--output", default=str(OUTPUT_PATH), help="Output JSON path")
    parser.add_argument("--skip-analysis", action="store_true",
                        help="Skip Claude deep dive analysis (faster)")
    parser.add_argument("--effective-grades",
                        default=str(DEFAULT_EG_CSV) if DEFAULT_EG_CSV.exists() else None,
                        help="Path to Student Progress Tracker CSV with effective grades")
    parser.add_argument("--s1-snapshot",
                        default=str(DEFAULT_S1_SNAPSHOT) if DEFAULT_S1_SNAPSHOT.exists() else None,
                        help="Path to S1 Snapshot Excel for S1 cohort identification")
    args = parser.parse_args()

    data = collect(args.csv, args.session, skip_analysis=args.skip_analysis,
                   effective_grades_csv=args.effective_grades,
                   s1_snapshot_path=args.s1_snapshot)

    output = Path(args.output)
    output.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("Dashboard data written to %s (%d students)", output, len(data["students"]))


if __name__ == "__main__":
    main()
